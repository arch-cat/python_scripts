import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


class HpsmScraper:
    def __init__(self, _driver_path):
        self.driver_path = driver_path
        self.driver = None

    def initialize_driver(self):
        options = Options()
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        # option.add_argument("--headless")  # Без запуска окна
        # option.add_argument("enable-features=NetworkServiceInProcess")
        # option.add_argument("disable-features=NetworkService")
        service = Service(executable_path=self.driver_path)
        self.driver = webdriver.Edge(service=service, options=options)

    def log_in(self, username, password):
        hpsm_username = self.driver.find_element(By.ID, "username")
        hpsm_username.send_keys(username)
        hpsm_password = self.driver.find_element(By.ID, "password")
        hpsm_password.send_keys(password)
        hpsm_login_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable((By.ID, "submitInter"))
        )
        hpsm_login_button.click()
        print("Успешный вход в систему")

    def scrape_data(self):
        # driver.maximize_window()
        # driver.set_page_load_timeout(30)
        self.driver.get("https://servicemanager.ru/hpsm/index.do?lang=ru")
        ci_list = ["123321", "321123", "321123"]

        # Begin
        self.log_in("{username}", "{password}")

        time.sleep(15)

        hpsm_search_button = self.driver.find_element(
            By.XPATH, "//button[@aria-label='Поиск']"
        )
        hpsm_search_button.click()
        time.sleep(1)

        adm_dict = {}
        for ci in ci_list:
            hpsm_search_input = self.driver.find_element(
                By.ID, "gs-trigger-cmdcombined"
            )
            hpsm_search_input.send_keys(f"{ci}")
            hpsm_search_input.send_keys(Keys.ENTER)

            time.sleep(8)

            self.driver.switch_to.frame(
                self.driver.find_elements(By.XPATH, f'//iframe[@title="КЭ: {ci}"]')[0]
            )
            hpsm_table = self.driver.find_element(By.XPATH, '//*[@id="X24_t"]')
            hpsm_table.click()

            rows = 1 + len(
                self.driver.find_elements(
                    By.XPATH,
                    f"/html/body/div[1]/div[1]/form/div[1]/div[3]/div/div[1]/fieldset/div[1]/div/div[2]/div["
                    f"2]/table/tbody/tr",
                )
            )
            list_name = f"{ci}"
            adm_dict[list_name] = []
            for r in range(1, rows):
                value = self.driver.find_element(
                    By.XPATH,
                    f"/html/body/div[1]/div[1]/form/div[1]/div[3]/div/div[1]/fieldset/div[1]/div/div[2]/div["
                    f"2]/table/tbody/tr[{r}]/td/div/div/div/div/div[1]/div[2]/input",
                )
                get_value = value.get_attribute("value")
                adm_dict[list_name].append(get_value)
                # print(get_value)
            self.driver.switch_to.default_content()
        print(adm_dict)

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        for col, key in enumerate(adm_dict.keys(), start=1):
            sheet.cell(row=row, column=col, value=key)
        for col, values in enumerate(adm_dict.values(), start=1):
            row = 2
            for value in values:
                sheet.cell(row=row, column=col, value=value)
                row += 1
        workbook.save("ci_list.xlsx")

        title = self.driver.title
        print(title)
        time.sleep(50)

    def close(self):
        if self.driver:
            self.driver.quit()


if __name__ == "__main__":
    driver_path = r"C:\Users\arch-cat\Downloads\edge\msedgedriver.exe"
    hpsm_scraper = HpsmScraper(driver_path)
    hpsm_scraper.initialize_driver()
    hpsm_scraper.scrape_data()
    hpsm_scraper.close()
